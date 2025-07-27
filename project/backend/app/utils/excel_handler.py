import pandas as pd
from typing import List, Dict, Any, Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Excel文件处理工具类
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def read_excel(self, file_content: bytes, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        读取Excel文件
        
        Args:
            file_content: Excel文件内容
            sheet_name: 工作表名称，如果为None则读取第一个工作表
        
        Returns:
            DataFrame: 读取的数据
        """
        try:
            if sheet_name:
                df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name)
            else:
                df = pd.read_excel(BytesIO(file_content))
            
            # 清理列名（去除空格和特殊字符）
            df.columns = df.columns.str.strip()
            
            return df
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            raise
    
    def create_excel(self, data: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> BytesIO:
        """
        创建Excel文件
        
        Args:
            data: 要写入的数据
            sheet_name: 工作表名称
        
        Returns:
            BytesIO: Excel文件内容
        """
        try:
            df = pd.DataFrame(data)
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表并设置样式
                worksheet = writer.sheets[sheet_name]
                self._apply_basic_styling(worksheet, df)
            
            output.seek(0)
            return output
        except Exception as e:
            logger.error(f"创建Excel文件失败: {e}")
            raise
    
    def create_styled_excel(self, data: List[Dict[str, Any]], 
                           sheet_name: str = "Sheet1",
                           title: str = None) -> BytesIO:
        """
        创建带样式的Excel文件
        
        Args:
            data: 要写入的数据
            sheet_name: 工作表名称
            title: 表格标题
        
        Returns:
            BytesIO: Excel文件内容
        """
        try:
            df = pd.DataFrame(data)
            output = BytesIO()
            
            # 创建工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            # 添加标题
            start_row = 1
            if title:
                worksheet.merge_cells(f'A1:{self._get_column_letter(len(df.columns))}1')
                title_cell = worksheet['A1']
                title_cell.value = title
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                title_cell.font = Font(size=16, bold=True, color='FFFFFF')
                worksheet.row_dimensions[1].height = 30
                start_row = 3
            
            # 添加数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 设置表头样式
                    if r_idx == start_row:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # 设置数据行样式
                        if r_idx % 2 == 0:
                            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 设置边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # 自动调整列宽
            self._auto_adjust_column_width(worksheet, df)
            
            # 保存到BytesIO
            workbook.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"创建带样式的Excel文件失败: {e}")
            raise
    
    def validate_excel_structure(self, df: pd.DataFrame, required_columns: List[str]) -> Dict[str, Any]:
        """
        验证Excel文件结构
        
        Args:
            df: 要验证的DataFrame
            required_columns: 必需的列名列表
        
        Returns:
            Dict: 验证结果
        """
        result = {
            'is_valid': True,
            'errors': [],
            'warnings': []
        }
        
        # 检查必需列
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            result['is_valid'] = False
            result['errors'].append(f"缺少必需列: {', '.join(missing_columns)}")
        
        # 检查空数据
        if df.empty:
            result['is_valid'] = False
            result['errors'].append("Excel文件中没有数据")
        
        # 检查重复列
        duplicate_columns = df.columns[df.columns.duplicated()].tolist()
        if duplicate_columns:
            result['warnings'].append(f"发现重复列: {', '.join(duplicate_columns)}")
        
        return result
    
    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        清理数据
        
        Args:
            df: 要清理的DataFrame
        
        Returns:
            DataFrame: 清理后的数据
        """
        # 去除完全空白的行
        df = df.dropna(how='all')
        
        # 去除列名中的空格
        df.columns = df.columns.str.strip()
        
        # 去除字符串列中的前后空格
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            # 将空字符串转换为NaN
            df[col] = df[col].replace('', None)
        
        return df
    
    def _apply_basic_styling(self, worksheet, df: pd.DataFrame):
        """
        应用基本样式
        """
        # 设置表头样式
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # 自动调整列宽
        self._auto_adjust_column_width(worksheet, df)
    
    def _auto_adjust_column_width(self, worksheet, df: pd.DataFrame):
        """
        自动调整列宽
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置最小和最大宽度
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_column_letter(self, column_number: int) -> str:
        """
        获取列字母
        """
        result = ""
        while column_number > 0:
            column_number -= 1
            result = chr(column_number % 26 + ord('A')) + result
            column_number //= 26
        return result
    
    @staticmethod
    def get_sheet_names(file_content: bytes) -> List[str]:
        """
        获取Excel文件中的所有工作表名称
        
        Args:
            file_content: Excel文件内容
        
        Returns:
            List[str]: 工作表名称列表
        """
        try:
            excel_file = pd.ExcelFile(BytesIO(file_content))
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"获取工作表名称失败: {e}")
            raise
    
    @staticmethod
    def convert_to_csv(file_content: bytes, sheet_name: Optional[str] = None) -> str:
        """
        将Excel文件转换为CSV格式
        
        Args:
            file_content: Excel文件内容
            sheet_name: 工作表名称
        
        Returns:
            str: CSV格式的数据
        """
        try:
            if sheet_name:
                df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name)
            else:
                df = pd.read_excel(BytesIO(file_content))
            
            return df.to_csv(index=False)
        except Exception as e:
            logger.error(f"转换为CSV失败: {e}")
            raise


# 创建全局实例
excel_handler = ExcelHandler()