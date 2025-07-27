"""报表导出工具

提供Excel、PDF、CSV等格式的报表导出功能
"""

import os
import io
from typing import Dict, Any, List
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端

class ReportExporter:
    """报表导出器"""
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_production_report_excel(self, data: Dict[str, Any], filename: str) -> str:
        """导出生产报表为Excel格式"""
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建概览工作表
        ws_overview = wb.create_sheet("生产概览")
        self._create_production_overview_sheet(ws_overview, data)
        
        # 创建车间统计工作表
        ws_workshop = wb.create_sheet("车间统计")
        self._create_workshop_stats_sheet(ws_workshop, data.get('workshop_stats', []))
        
        # 创建日产量工作表
        ws_daily = wb.create_sheet("日产量统计")
        self._create_daily_production_sheet(ws_daily, data.get('daily_production', []))
        
        # 保存文件
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_quality_report_excel(self, data: Dict[str, Any], filename: str) -> str:
        """导出质量报表为Excel格式"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # 创建概览工作表
        ws_overview = wb.create_sheet("质量概览")
        self._create_quality_overview_sheet(ws_overview, data)
        
        # 创建缺陷分析工作表
        ws_defect = wb.create_sheet("缺陷分析")
        self._create_defect_analysis_sheet(ws_defect, data.get('defect_analysis', []))
        
        # 创建产品质量工作表
        ws_product = wb.create_sheet("产品质量")
        self._create_product_quality_sheet(ws_product, data.get('product_quality', []))
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_equipment_report_excel(self, data: Dict[str, Any], filename: str) -> str:
        """导出设备报表为Excel格式"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # 创建概览工作表
        ws_overview = wb.create_sheet("设备概览")
        self._create_equipment_overview_sheet(ws_overview, data)
        
        # 创建维护统计工作表
        ws_maintenance = wb.create_sheet("维护统计")
        self._create_maintenance_stats_sheet(ws_maintenance, data.get('maintenance_stats', []))
        
        # 创建故障分析工作表
        ws_fault = wb.create_sheet("故障分析")
        self._create_fault_analysis_sheet(ws_fault, data.get('fault_analysis', []))
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_to_pdf(self, data: Dict[str, Any], report_type: str, filename: str) -> str:
        """导出报表为PDF格式"""
        filepath = os.path.join(self.output_dir, filename)
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # 居中
        )
        
        if report_type == 'production':
            story.append(Paragraph("生产报表", title_style))
            self._add_production_pdf_content(story, data, styles)
        elif report_type == 'quality':
            story.append(Paragraph("质量报表", title_style))
            self._add_quality_pdf_content(story, data, styles)
        elif report_type == 'equipment':
            story.append(Paragraph("设备报表", title_style))
            self._add_equipment_pdf_content(story, data, styles)
        
        doc.build(story)
        return filepath
    
    def export_to_csv(self, data: Dict[str, Any], report_type: str, filename: str) -> str:
        """导出报表为CSV格式"""
        filepath = os.path.join(self.output_dir, filename)
        
        if report_type == 'production':
            df = pd.DataFrame(data.get('workshop_stats', []))
        elif report_type == 'quality':
            df = pd.DataFrame(data.get('product_quality', []))
        elif report_type == 'equipment':
            df = pd.DataFrame(data.get('maintenance_stats', []))
        else:
            df = pd.DataFrame()
        
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        return filepath
    
    def _create_production_overview_sheet(self, ws, data):
        """创建生产概览工作表"""
        # 设置标题
        ws['A1'] = '生产报表概览'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 基础统计
        headers = ['指标', '数值']
        ws.append([])
        ws.append(headers)
        
        stats = [
            ['总计划数', data.get('total_plans', 0)],
            ['已完成计划', data.get('completed_plans', 0)],
            ['进行中计划', data.get('in_progress_plans', 0)],
            ['逾期计划', data.get('overdue_plans', 0)],
            ['完成率(%)', f"{data.get('completion_rate', 0):.1f}"],
            ['平均进度(%)', f"{data.get('avg_progress', 0):.1f}"]
        ]
        
        for stat in stats:
            ws.append(stat)
    
    def _create_workshop_stats_sheet(self, ws, workshop_stats):
        """创建车间统计工作表"""
        ws['A1'] = '车间统计'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['车间', '计划数', '平均进度(%)']
        ws.append([])
        ws.append(headers)
        
        for stat in workshop_stats:
            ws.append([
                stat.get('workshop', ''),
                stat.get('total_plans', 0),
                f"{stat.get('avg_progress', 0):.1f}"
            ])
    
    def _create_daily_production_sheet(self, ws, daily_production):
        """创建日产量统计工作表"""
        ws['A1'] = '日产量统计'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['日期', '产量']
        ws.append([])
        ws.append(headers)
        
        for daily in daily_production:
            ws.append([
                daily.get('date', ''),
                daily.get('count', 0)
            ])
    
    def _create_quality_overview_sheet(self, ws, data):
        """创建质量概览工作表"""
        ws['A1'] = '质量报表概览'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        headers = ['指标', '数值']
        ws.append([])
        ws.append(headers)
        
        stats = [
            ['总检查次数', data.get('total_checks', 0)],
            ['通过次数', data.get('passed_checks', 0)],
            ['未通过次数', data.get('failed_checks', 0)],
            ['合格率(%)', f"{data.get('pass_rate', 0):.1f}"],
            ['缺陷率(%)', f"{data.get('defect_rate', 0):.1f}"]
        ]
        
        for stat in stats:
            ws.append(stat)
    
    def _create_defect_analysis_sheet(self, ws, defect_analysis):
        """创建缺陷分析工作表"""
        ws['A1'] = '缺陷分析'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['缺陷类型', '数量', '占比(%)']
        ws.append([])
        ws.append(headers)
        
        for defect in defect_analysis:
            ws.append([
                defect.get('defect_type', ''),
                defect.get('count', 0),
                f"{defect.get('percentage', 0):.1f}"
            ])
    
    def _create_product_quality_sheet(self, ws, product_quality):
        """创建产品质量工作表"""
        ws['A1'] = '产品质量统计'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['产品名称', '检查次数', '通过次数', '合格率(%)']
        ws.append([])
        ws.append(headers)
        
        for product in product_quality:
            ws.append([
                product.get('product_name', ''),
                product.get('total_checks', 0),
                product.get('passed_checks', 0),
                f"{product.get('pass_rate', 0):.1f}"
            ])
    
    def _create_equipment_overview_sheet(self, ws, data):
        """创建设备概览工作表"""
        ws['A1'] = '设备报表概览'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        headers = ['指标', '数值']
        ws.append([])
        ws.append(headers)
        
        stats = [
            ['设备总数', data.get('total_equipment', 0)],
            ['运行中设备', data.get('running_equipment', 0)],
            ['维护中设备', data.get('maintenance_equipment', 0)],
            ['故障设备', data.get('fault_equipment', 0)],
            ['设备利用率(%)', f"{data.get('utilization_rate', 0):.1f}"]
        ]
        
        for stat in stats:
            ws.append(stat)
    
    def _create_maintenance_stats_sheet(self, ws, maintenance_stats):
        """创建维护统计工作表"""
        ws['A1'] = '维护统计'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['维护类型', '次数', '平均成本']
        ws.append([])
        ws.append(headers)
        
        for maintenance in maintenance_stats:
            ws.append([
                maintenance.get('maintenance_type', ''),
                maintenance.get('count', 0),
                f"¥{maintenance.get('avg_cost', 0):.0f}"
            ])
    
    def _create_fault_analysis_sheet(self, ws, fault_analysis):
        """创建故障分析工作表"""
        ws['A1'] = '故障分析'
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = ['设备类型', '故障次数', '故障率(%)']
        ws.append([])
        ws.append(headers)
        
        for fault in fault_analysis:
            ws.append([
                fault.get('equipment_type', ''),
                fault.get('fault_count', 0),
                f"{fault.get('fault_rate', 0):.2f}"
            ])
    
    def _add_production_pdf_content(self, story, data, styles):
        """添加生产报表PDF内容"""
        # 基础统计表格
        stats_data = [
            ['指标', '数值'],
            ['总计划数', str(data.get('total_plans', 0))],
            ['已完成计划', str(data.get('completed_plans', 0))],
            ['进行中计划', str(data.get('in_progress_plans', 0))],
            ['逾期计划', str(data.get('overdue_plans', 0))],
            ['完成率', f"{data.get('completion_rate', 0):.1f}%"],
            ['平均进度', f"{data.get('avg_progress', 0):.1f}%"]
        ]
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
    
    def _add_quality_pdf_content(self, story, data, styles):
        """添加质量报表PDF内容"""
        stats_data = [
            ['指标', '数值'],
            ['总检查次数', str(data.get('total_checks', 0))],
            ['通过次数', str(data.get('passed_checks', 0))],
            ['未通过次数', str(data.get('failed_checks', 0))],
            ['合格率', f"{data.get('pass_rate', 0):.1f}%"],
            ['缺陷率', f"{data.get('defect_rate', 0):.1f}%"]
        ]
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
    
    def _add_equipment_pdf_content(self, story, data, styles):
        """添加设备报表PDF内容"""
        stats_data = [
            ['指标', '数值'],
            ['设备总数', str(data.get('total_equipment', 0))],
            ['运行中设备', str(data.get('running_equipment', 0))],
            ['维护中设备', str(data.get('maintenance_equipment', 0))],
            ['故障设备', str(data.get('fault_equipment', 0))],
            ['设备利用率', f"{data.get('utilization_rate', 0):.1f}%"]
        ]
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))

# 全局导出器实例
report_exporter = ReportExporter()