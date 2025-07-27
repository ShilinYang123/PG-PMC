#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件处理服务

提供文件上传、下载、Excel导入导出等功能：
- 文件上传处理
- Excel文件读取和写入
- 文件格式验证
- 文件存储管理
"""

import os
import uuid
import shutil
from typing import List, Dict, Any, Optional, Union, BinaryIO
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import UploadFile, HTTPException
from loguru import logger
import mimetypes
import hashlib

from ..core.config import settings
from ..core.exceptions import ValidationException, BusinessException


class FileService:
    """文件处理服务类"""
    
    def __init__(self):
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.temp_dir = self.upload_dir / "temp"
        self.export_dir = self.upload_dir / "exports"
        
        # 确保目录存在
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        # 支持的文件类型
        self.allowed_extensions = {
            '.xlsx', '.xls', '.csv', '.txt',
            '.pdf', '.doc', '.docx',
            '.jpg', '.jpeg', '.png', '.gif'
        }
        
        # MIME类型映射
        self.mime_types = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
            'application/vnd.ms-excel': '.xls',
            'text/csv': '.csv',
            'text/plain': '.txt',
            'application/pdf': '.pdf',
            'image/jpeg': '.jpg',
            'image/png': '.png'
        }
    
    async def upload_file(
        self, 
        file: UploadFile, 
        category: str = "general",
        max_size: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        上传文件
        
        Args:
            file: 上传的文件
            category: 文件分类
            max_size: 最大文件大小（字节）
            
        Returns:
            Dict: 文件信息
        """
        try:
            # 验证文件
            await self._validate_file(file, max_size)
            
            # 生成文件名
            file_id = str(uuid.uuid4())
            original_name = file.filename or "unknown"
            file_ext = Path(original_name).suffix.lower()
            
            if not file_ext:
                # 根据MIME类型推断扩展名
                file_ext = self.mime_types.get(file.content_type, '.bin')
            
            filename = f"{file_id}{file_ext}"
            
            # 创建分类目录
            category_dir = self.upload_dir / category
            category_dir.mkdir(parents=True, exist_ok=True)
            
            # 保存文件
            file_path = category_dir / filename
            content = await file.read()
            
            with open(file_path, "wb") as f:
                f.write(content)
            
            # 计算文件哈希
            file_hash = hashlib.md5(content).hexdigest()
            
            # 获取文件信息
            file_info = {
                "file_id": file_id,
                "original_name": original_name,
                "filename": filename,
                "file_path": str(file_path),
                "file_url": f"/api/files/{category}/{filename}",
                "file_size": len(content),
                "file_type": file.content_type,
                "file_ext": file_ext,
                "file_hash": file_hash,
                "category": category,
                "upload_time": datetime.now(),
                "status": "uploaded"
            }
            
            logger.info(f"文件上传成功: {original_name} -> {filename}")
            return file_info
            
        except Exception as e:
            logger.error(f"文件上传失败: {e}")
            raise BusinessException(f"文件上传失败: {str(e)}")
    
    async def _validate_file(
        self, 
        file: UploadFile, 
        max_size: Optional[int] = None
    ) -> None:
        """
        验证文件
        
        Args:
            file: 上传的文件
            max_size: 最大文件大小
        """
        # 检查文件名
        if not file.filename:
            raise ValidationException("文件名不能为空")
        
        # 检查文件扩展名
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in self.allowed_extensions:
            raise ValidationException(
                f"不支持的文件类型: {file_ext}。"
                f"支持的类型: {', '.join(self.allowed_extensions)}"
            )
        
        # 检查文件大小
        max_file_size = max_size or settings.MAX_FILE_SIZE
        if file.size and file.size > max_file_size:
            raise ValidationException(
                f"文件大小超过限制: {file.size} > {max_file_size} 字节"
            )
    
    def read_excel(
        self, 
        file_path: Union[str, Path], 
        sheet_name: Optional[str] = None,
        header_row: int = 0,
        max_rows: Optional[int] = None
    ) -> pd.DataFrame:
        """
        读取Excel文件
        
        Args:
            file_path: 文件路径
            sheet_name: 工作表名称
            header_row: 标题行索引
            max_rows: 最大读取行数
            
        Returns:
            DataFrame: 数据
        """
        try:
            file_path = Path(file_path)
            
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            # 读取Excel文件
            if file_path.suffix.lower() == '.csv':
                df = pd.read_csv(
                    file_path, 
                    header=header_row,
                    nrows=max_rows,
                    encoding='utf-8-sig'
                )
            else:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=header_row,
                    nrows=max_rows
                )
            
            # 清理数据
            df = df.dropna(how='all')  # 删除全空行
            df.columns = df.columns.astype(str)  # 确保列名为字符串
            
            logger.info(f"成功读取Excel文件: {file_path}, 行数: {len(df)}")
            return df
            
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            raise BusinessException(f"读取Excel文件失败: {str(e)}")
    
    def write_excel(
        self, 
        data: Union[pd.DataFrame, List[Dict[str, Any]]], 
        file_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        include_index: bool = False,
        style_config: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        写入Excel文件
        
        Args:
            data: 数据
            file_path: 文件路径
            sheet_name: 工作表名称
            include_index: 是否包含索引
            style_config: 样式配置
            
        Returns:
            str: 文件路径
        """
        try:
            file_path = Path(file_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 转换数据格式
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data.copy()
            
            # 写入Excel文件
            with pd.ExcelWriter(
                file_path, 
                engine='openpyxl',
                options={'remove_timezone': True}
            ) as writer:
                df.to_excel(
                    writer, 
                    sheet_name=sheet_name, 
                    index=include_index
                )
                
                # 应用样式
                if style_config:
                    self._apply_excel_style(
                        writer.book[sheet_name], 
                        style_config
                    )
            
            logger.info(f"成功写入Excel文件: {file_path}, 行数: {len(df)}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"写入Excel文件失败: {e}")
            raise BusinessException(f"写入Excel文件失败: {str(e)}")
    
    def _apply_excel_style(
        self, 
        worksheet, 
        style_config: Dict[str, Any]
    ) -> None:
        """
        应用Excel样式
        
        Args:
            worksheet: 工作表对象
            style_config: 样式配置
        """
        try:
            # 标题样式
            if style_config.get('header_style'):
                header_style = style_config['header_style']
                for cell in worksheet[1]:
                    if header_style.get('font'):
                        cell.font = Font(**header_style['font'])
                    if header_style.get('fill'):
                        cell.fill = PatternFill(**header_style['fill'])
                    if header_style.get('alignment'):
                        cell.alignment = Alignment(**header_style['alignment'])
            
            # 数据样式
            if style_config.get('data_style'):
                data_style = style_config['data_style']
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if data_style.get('alignment'):
                            cell.alignment = Alignment(**data_style['alignment'])
            
            # 边框样式
            if style_config.get('border'):
                border = Border(**style_config['border'])
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = border
            
            # 自动调整列宽
            if style_config.get('auto_width', True):
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
        except Exception as e:
            logger.warning(f"应用Excel样式失败: {e}")
    
    def generate_export_file(
        self, 
        data: Union[pd.DataFrame, List[Dict[str, Any]]], 
        filename: str,
        file_format: str = "xlsx",
        style_config: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        生成导出文件
        
        Args:
            data: 数据
            filename: 文件名
            file_format: 文件格式
            style_config: 样式配置
            
        Returns:
            Dict: 文件信息
        """
        try:
            # 生成文件路径
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_id = str(uuid.uuid4())[:8]
            
            if not filename.endswith(f'.{file_format}'):
                filename = f"{filename}_{timestamp}_{file_id}.{file_format}"
            
            file_path = self.export_dir / filename
            
            # 根据格式导出文件
            if file_format.lower() in ['xlsx', 'xls']:
                # 默认样式配置
                default_style = {
                    'header_style': {
                        'font': {'bold': True, 'color': 'FFFFFF'},
                        'fill': {'start_color': '366092', 'end_color': '366092', 'fill_type': 'solid'},
                        'alignment': {'horizontal': 'center', 'vertical': 'center'}
                    },
                    'data_style': {
                        'alignment': {'horizontal': 'left', 'vertical': 'center'}
                    },
                    'auto_width': True
                }
                
                if style_config:
                    default_style.update(style_config)
                
                self.write_excel(
                    data, 
                    file_path, 
                    style_config=default_style
                )
                
            elif file_format.lower() == 'csv':
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                else:
                    df = data
                
                df.to_csv(
                    file_path, 
                    index=False, 
                    encoding='utf-8-sig'
                )
            
            else:
                raise ValidationException(f"不支持的导出格式: {file_format}")
            
            # 生成文件信息
            file_info = {
                "file_id": file_id,
                "filename": filename,
                "file_path": str(file_path),
                "file_url": f"/api/files/exports/{filename}",
                "file_size": file_path.stat().st_size,
                "file_format": file_format,
                "created_time": datetime.now(),
                "status": "ready"
            }
            
            logger.info(f"成功生成导出文件: {filename}")
            return file_info
            
        except Exception as e:
            logger.error(f"生成导出文件失败: {e}")
            raise BusinessException(f"生成导出文件失败: {str(e)}")
    
    def delete_file(self, file_path: Union[str, Path]) -> bool:
        """
        删除文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            file_path = Path(file_path)
            if file_path.exists():
                file_path.unlink()
                logger.info(f"成功删除文件: {file_path}")
                return True
            return False
            
        except Exception as e:
            logger.error(f"删除文件失败: {e}")
            return False
    
    def cleanup_temp_files(self, max_age_hours: int = 24) -> int:
        """
        清理临时文件
        
        Args:
            max_age_hours: 最大保留时间（小时）
            
        Returns:
            int: 清理的文件数量
        """
        try:
            count = 0
            cutoff_time = datetime.now().timestamp() - (max_age_hours * 3600)
            
            for file_path in self.temp_dir.iterdir():
                if file_path.is_file():
                    if file_path.stat().st_mtime < cutoff_time:
                        file_path.unlink()
                        count += 1
            
            logger.info(f"清理临时文件完成，删除 {count} 个文件")
            return count
            
        except Exception as e:
            logger.error(f"清理临时文件失败: {e}")
            return 0


# 全局文件服务实例
file_service = FileService()


# 导出主要接口
__all__ = [
    'FileService',
    'file_service'
]